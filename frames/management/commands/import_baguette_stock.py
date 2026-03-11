from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from frames.models import Baguette


class Command(BaseCommand):
    help = "Импорт остатков багетов из xlsx (Багет -> name, остаток -> stock_quantity)"

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            type=str,
            required=True,
            help="Путь к xlsx файлу с остатками багетов",
        )

    @staticmethod
    def _to_decimal(value):
        if value in (None, ""):
            return Decimal("0")

        normalized = str(value).strip().replace(",", ".")
        if normalized.endswith("."):
            normalized = normalized[:-1]

        try:
            return Decimal(normalized)
        except (InvalidOperation, ValueError):
            return Decimal("0")

    def handle(self, *args, **options):
        file_path = options["file"]
        wb = load_workbook(file_path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        created_count = 0
        updated_count = 0
        skipped_count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            name_raw = row[1] if len(row) > 1 else None
            stock_raw = row[4] if len(row) > 4 else None

            if not name_raw:
                skipped_count += 1
                continue

            name = str(name_raw).strip()
            if not name:
                skipped_count += 1
                continue

            stock_quantity = self._to_decimal(stock_raw)

            _, created = Baguette.objects.update_or_create(
                name=name,
                defaults={
                    "stock_quantity": stock_quantity,
                    "price": Decimal("0"),
                    "width": Decimal("0"),
                },
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        self.stdout.write(
            self.style.SUCCESS(
                "Импорт завершен: "
                f"создано={created_count}, "
                f"обновлено={updated_count}, "
                f"пропущено={skipped_count}"
            )
        )
